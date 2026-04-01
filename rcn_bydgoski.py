
"""
RCN – Powiat Bydgoski (0403) | Parser cen i parametrów transakcji
==================================================================
Problem: GeoPandas/GDAL przy czytaniu GML pomija zagnieżdżone atrybuty XML
         (cenaTransakcjiBrutto, powierzchnia, rodzajNieruchomosci itp.)
Rozwiązanie: parsowanie GML bezpośrednio przez lxml + ElementTree

Wynik: rcn_bydgoski_ceny.xlsx z arkuszami:
  - Lokale    (mieszkania, lokale użytkowe) – cena, m2, cena/m2, adres
  - Działki   (grunty)
  - Budynki   (domy)
  - Podsumowanie
"""

import requests
import zipfile
import io
import re
from pathlib import Path
from lxml import etree
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

TERYT       = "0403"
OUTPUT_PATH = Path("rcn_bydgoski_ceny.xlsx")
GEOFORUM_URL = f"https://rcn.geoforum.pl/download.php?teryt={TERYT}"

# ─── Przestrzenie nazw GML/RCN ───────────────────────────────────────────────
# GUGiK używa schematu rcn: – namespace może się różnić między powiatami,
# dlatego wykrywamy go dynamicznie z pliku GML.

def wykryj_namespace(root):
    """Wykrywa namespace rcn: z korzenia dokumentu GML."""
    ns = {}
    for prefix, uri in root.nsmap.items():
        if prefix:
            ns[prefix] = uri
        else:
            ns["default"] = uri
    # Szukamy namespace zawierającego 'rcn' lub 'RCN'
    rcn_ns = next((uri for uri in ns.values() if "rcn" in uri.lower()), None)
    gml_ns = next((uri for uri in ns.values() if "opengis.net/gml" in uri), None)
    return ns, rcn_ns, gml_ns


# ─── Mapowanie tagów → kolumny Excel ─────────────────────────────────────────
# Pełna lista atrybutów ze schematu XSD RCN (rozporządzenie 2021)

ATRYBUTY_TRANSAKCJA = {
    "cenaTransakcjiBrutto":       "cena_brutto_PLN",
    "dataAktuNotarialnego":       "data_aktu",
    "rodzajTransakcji":           "rodzaj_transakcji",
    "stronaSprzedajaca":          "strona_sprzedajaca",
    "stronaKupujaca":             "strona_kupujaca",
    "rodzajRynku":                "rodzaj_rynku",       # 1=pierwotny, 2=wtórny
    "oznaczenieDokumentu":        "nr_aktu",
    "warunekZawarcia":            "warunek_zawarcia",
}

ATRYBUTY_LOKAL = {
    "powierzchniaUzytkowa":       "powierzchnia_m2",
    "liczbaPokoi":                "liczba_pokoi",
    "kondygnacja":                "kondygnacja",
    "liczbaKondygnacji":          "liczba_kondygnacji_budynku",
    "rokBudowy":                  "rok_budowy",
    "adres":                      "adres_raw",
    "numerLokalu":                "nr_lokalu",
    "rodzajLokalu":               "rodzaj_lokalu",     # mieszkalny/użytkowy
    "udzialWNieruchomosci":       "udzial",
}

ATRYBUTY_DZIALKA = {
    "powierzchnia":               "powierzchnia_m2",
    "numerEwidencyjny":           "nr_ewidencyjny",
    "rodzajUzytku":               "rodzaj_uzytku",
    "klasaGruntuDominujaca":      "klasa_gruntu",
    "przeznaczenieMPZP":          "mpzp_przeznaczenie",
    "uzbrojenie":                 "uzbrojenie",
    "adres":                      "adres_raw",
}

ATRYBUTY_BUDYNEK = {
    "powierzchniaUzytkowa":       "powierzchnia_uzytkowa_m2",
    "powierzchniaZabudowy":       "powierzchnia_zabudowy_m2",
    "rokBudowy":                  "rok_budowy",
    "rodzajBudynku":              "rodzaj_budynku",
    "liczbaPokoi":                "liczba_pokoi",
    "liczbaKondygnacji":          "liczba_kondygnacji",
    "adres":                      "adres_raw",
}

SLOWNIKI = {
    "rodzaj_rynku": {"1": "pierwotny", "2": "wtórny"},
    "rodzaj_transakcji": {
        "1": "sprzedaż", "2": "zamiana", "3": "darowizna",
        "4": "inne odpłatne", "5": "inne nieodpłatne"
    },
    "strona_sprzedajaca": {
        "1": "os. fizyczna", "2": "os. prawna", "3": "j. sam. teryt.",
        "4": "Skarb Państwa", "5": "inna"
    },
}


# ─── Parser GML ──────────────────────────────────────────────────────────────

def tekst(element, tag, ns_uri):
    """Bezpieczne pobranie tekstu zagnieżdżonego tagu (obsługuje różne ns)."""
    # próba z podanym ns
    el = element.find(f"{{{ns_uri}}}{tag}") if ns_uri else None
    if el is None:
        # szukaj w całym drzewie bez ns
        el = element.find(f".//*[local-name()='{tag}']")
    return el.text.strip() if el is not None and el.text else None


def parsuj_adres(element):
    """Wyciąga adres z zagnieżdżonego obiektu adresowego."""
    parts = []
    for tag in ["miejscowosc", "ulica", "numerPorzadkowy", "kodPocztowy"]:
        val = element.find(f".//*[local-name()='{tag}']")
        if val is not None and val.text:
            parts.append(val.text.strip())
    return ", ".join(parts) if parts else None


def parsuj_wspolrzedne(element):
    """Wyciąga lat/lon z Point lub centroidu geometrii."""
    pos = element.find(".//*[local-name()='pos']")
    if pos is not None and pos.text:
        coords = pos.text.strip().split()
        if len(coords) >= 2:
            return float(coords[0]), float(coords[1])
    return None, None


def parsuj_gml(gml_bytes: bytes) -> list[dict]:
    """
    Parsuje plik GML RCN i zwraca listę słowników z wszystkimi atrybutami.
    Obsługuje zagnieżdżone obiekty (Lokal, Dzialka, Budynek wewnątrz Transakcji).
    """
    try:
        root = etree.fromstring(gml_bytes)
    except etree.XMLSyntaxError as e:
        print(f"  ✗ Błąd XML: {e}")
        return []

    ns, rcn_ns, gml_ns = wykryj_namespace(root)
    rekordy = []

    # Iteruj po wszystkich obiektach Feature w dokumencie
    for member in root.iter():
        local = etree.QName(member.tag).localname if member.tag else ""

        # Szukamy głównych obiektów transakcji/lokalu/działki/budynku
        if local not in ("Transakcja", "ZbycieLokalu", "ZbycieDzialki",
                         "ZbycieBudynku", "PrzedmiotTransakcji",
                         "Lokal", "Dzialka", "Budynek"):
            continue

        rekord = {"_typ": local}

        # ── Atrybuty transakcji ──
        for xml_tag, col_name in ATRYBUTY_TRANSAKCJA.items():
            val = element_tekst_lokalny(member, xml_tag)
            if val:
                rekord[col_name] = val

        # ── Atrybuty lokalu ──
        if local in ("Lokal", "ZbycieLokalu"):
            for xml_tag, col_name in ATRYBUTY_LOKAL.items():
                val = element_tekst_lokalny(member, xml_tag)
                if val:
                    rekord[col_name] = val

        # ── Atrybuty działki ──
        elif local in ("Dzialka", "ZbycieDzialki"):
            for xml_tag, col_name in ATRYBUTY_DZIALKA.items():
                val = element_tekst_lokalny(member, xml_tag)
                if val:
                    rekord[col_name] = val

        # ── Atrybuty budynku ──
        elif local in ("Budynek", "ZbycieBudynku"):
            for xml_tag, col_name in ATRYBUTY_BUDYNEK.items():
                val = element_tekst_lokalny(member, xml_tag)
                if val:
                    rekord[col_name] = val

        # ── Adres (zagnieżdżony) ──
        adres = parsuj_adres(member)
        if adres:
            rekord["adres"] = adres

        # ── Współrzędne ──
        lat, lon = parsuj_wspolrzedne(member)
        if lat:
            rekord["lat"] = lat
            rekord["lon"] = lon

        if len(rekord) > 1:  # ma coś poza _typ
            rekordy.append(rekord)

    return rekordy


def element_tekst_lokalny(parent, tag_local):
    """Szuka tagu po local-name() ignorując namespace."""
    el = parent.find(f".//*[local-name()='{tag_local}']")
    # Jeśli znalazł element z dziećmi (np. adres) – pomiń, bo ma osobny parser
    if el is not None and el.text and el.text.strip():
        return el.text.strip()
    return None


# ─── Pobieranie ZIP ───────────────────────────────────────────────────────────

def pobierz_zip() -> dict[str, bytes]:
    """Pobiera ZIP z geoforum i zwraca {nazwa_pliku: bytes} dla plików GML."""
    print(f"Pobieranie danych RCN dla powiatu {TERYT}...")
    headers = {"User-Agent": "Mozilla/5.0 (rcn-parser/2.0; research)"}
    r = requests.get(GEOFORUM_URL, headers=headers, timeout=180, stream=True)

    if r.status_code != 200:
        raise RuntimeError(f"HTTP {r.status_code} – nie można pobrać ZIP")

    wynik = {}
    with zipfile.ZipFile(io.BytesIO(r.content)) as zf:
        pliki = zf.namelist()
        print(f"  ZIP zawiera {len(pliki)} plików: {pliki}")
        for nazwa in pliki:
            if nazwa.lower().endswith(".gml"):
                with zf.open(nazwa) as f:
                    wynik[Path(nazwa).stem] = f.read()
    return wynik


# ─── Post-processing ──────────────────────────────────────────────────────────

def przetworz_df(df: pd.DataFrame, typ: str) -> pd.DataFrame:
    """Konwersja typów, słowniki, kolumna cena/m2."""

    # Konwersja numeryczna
    for col in ["cena_brutto_PLN", "powierzchnia_m2", "powierzchnia_uzytkowa_m2",
                "powierzchnia_zabudowy_m2", "liczba_pokoi", "kondygnacja",
                "rok_budowy", "liczba_kondygnacji", "liczba_kondygnacji_budynku"]:
        if col in df.columns:
            df[col] = pd.to_numeric(df[col], errors="coerce")

    # Słowniki kodów
    for col, slownik in SLOWNIKI.items():
        if col in df.columns:
            df[col] = df[col].astype(str).map(slownik).fillna(df[col])

    # Cena za m2 (tylko gdy są obie kolumny)
    pow_col = "powierzchnia_m2" if "powierzchnia_m2" in df.columns else \
              "powierzchnia_uzytkowa_m2" if "powierzchnia_uzytkowa_m2" in df.columns else None
    if "cena_brutto_PLN" in df.columns and pow_col:
        df["cena_za_m2_PLN"] = (df["cena_brutto_PLN"] / df[pow_col]).round(0)
        df["cena_za_m2_PLN"] = df["cena_za_m2_PLN"].replace([float("inf"), float("-inf")], pd.NA)

    # Usuń kolumnę _typ
    df = df.drop(columns=["_typ"], errors="ignore")

    # Posortuj po dacie
    if "data_aktu" in df.columns:
        df = df.sort_values("data_aktu", ascending=False)

    return df


# ─── Excel ────────────────────────────────────────────────────────────────────

HEADER_FILL = PatternFill("solid", start_color="1F4E79")
ALT_FILL    = PatternFill("solid", start_color="D6E4F0")
TITLE_FILL  = PatternFill("solid", start_color="2E75B6")
PRICE_FILL  = PatternFill("solid", start_color="E2EFDA")   # zielonkawy dla cen
THIN        = Side(style="thin", color="BFBFBF")
BORDER      = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

KLUCZOWE_KOLUMNY = [
    "data_aktu", "cena_brutto_PLN", "powierzchnia_m2",
    "powierzchnia_uzytkowa_m2", "cena_za_m2_PLN",
    "liczba_pokoi", "kondygnacja", "rok_budowy",
    "rodzaj_rynku", "rodzaj_transakcji", "adres",
]

def fmt_header(cell):
    cell.font      = Font(name="Arial", bold=True, color="FFFFFF", size=10)
    cell.fill      = HEADER_FILL
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    cell.border    = BORDER

def fmt_title(ws, row, text, n_cols):
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=max(n_cols, 1))
    c = ws.cell(row=row, column=1, value=text)
    c.font      = Font(name="Arial", bold=True, color="FFFFFF", size=13)
    c.fill      = TITLE_FILL
    c.alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[row].height = 28

def write_df_to_sheet(ws, df, title):
    if df is None or len(df) == 0:
        ws.cell(row=1, column=1, value="Brak danych")
        return

    # Posortuj kolumny: kluczowe pierwsze, reszta alfabetycznie
    kluczowe = [c for c in KLUCZOWE_KOLUMNY if c in df.columns]
    pozostale = sorted([c for c in df.columns if c not in kluczowe])
    kolumny = kluczowe + pozostale
    df = df[kolumny]

    fmt_title(ws, 1, title, len(kolumny))

    # Nagłówki (wiersz 2)
    for ci, col in enumerate(kolumny, 1):
        fmt_header(ws.cell(row=2, column=ci, value=col))
    ws.row_dimensions[2].height = 30
    ws.freeze_panes = "A3"

    # Dane (od wiersza 3)
    cena_cols = {ci+1 for ci, c in enumerate(kolumny)
                 if "cena" in c or "powierzchnia" in c}

    for ri, (_, row_data) in enumerate(df.iterrows(), 3):
        fill = ALT_FILL if ri % 2 == 0 else None
        for ci, val in enumerate(row_data, 1):
            c = ws.cell(row=ri, column=ci, value=val)
            c.border    = BORDER
            c.font      = Font(name="Arial", size=9)
            c.alignment = Alignment(vertical="center")
            if ci in cena_cols and pd.notna(val):
                c.fill      = PRICE_FILL
                c.font      = Font(name="Arial", size=9, bold=True, color="1F4E79")
                if isinstance(val, (int, float)):
                    c.number_format = '#,##0'
            elif fill:
                c.fill = fill

    # Szerokości
    for ci, col in enumerate(kolumny, 1):
        try:
            max_len = max(len(str(col)),
                         df[col].dropna().astype(str).str.len().max() if len(df) > 0 else 0)
        except Exception:
            max_len = len(str(col))
        ws.column_dimensions[get_column_letter(ci)].width = min(max_len + 3, 45)

    # Wiersz statystyk
    last = 2 + len(df)
    ws.cell(row=last + 2, column=1, value="Liczba rekordów:").font = Font(bold=True, name="Arial")
    ws.cell(row=last + 2, column=2, value=f"=COUNTA(A3:A{last})").font = \
        Font(bold=True, name="Arial", color="1F4E79")

    if "cena_brutto_PLN" in kolumny:
        cc = kolumny.index("cena_brutto_PLN") + 1
        col_ltr = get_column_letter(cc)
        ws.cell(row=last + 3, column=1, value="Mediana ceny [PLN]:").font = Font(bold=True, name="Arial")
        ws.cell(row=last + 3, column=2,
                value=f"=MEDIAN({col_ltr}3:{col_ltr}{last})").font = \
            Font(bold=True, name="Arial", color="1F4E79")
        ws.cell(row=last + 3, column=2).number_format = '#,##0'

        if "cena_za_m2_PLN" in kolumny:
            cm2 = kolumny.index("cena_za_m2_PLN") + 1
            cm2_ltr = get_column_letter(cm2)
            ws.cell(row=last + 4, column=1, value="Mediana ceny/m² [PLN]:").font = Font(bold=True, name="Arial")
            ws.cell(row=last + 4, column=2,
                    value=f"=MEDIAN({cm2_ltr}3:{cm2_ltr}{last})").font = \
                Font(bold=True, name="Arial", color="1F4E79")
            ws.cell(row=last + 4, column=2).number_format = '#,##0'

    ws.sheet_view.showGridLines = False


def write_summary(ws, info_list):
    ws.sheet_view.showGridLines = False
    fmt_title(ws, 1, f"RCN Powiat Bydgoski (0403) | Podsumowanie", 5)
    ws.row_dimensions[1].height = 28

    hdrs = ["Typ", "Rekordów", "Z ceną", "Mediana ceny PLN", "Mediana ceny/m²"]
    for ci, h in enumerate(hdrs, 1):
        fmt_header(ws.cell(row=3, column=ci, value=h))

    for ri, info in enumerate(info_list, 4):
        fill = ALT_FILL if ri % 2 == 0 else None
        for ci, val in enumerate(info, 1):
            c = ws.cell(row=ri, column=ci, value=val)
            c.font   = Font(name="Arial", size=10)
            c.border = BORDER
            if fill:
                c.fill = fill

    for ci, w in enumerate([22, 14, 12, 22, 22], 1):
        ws.column_dimensions[get_column_letter(ci)].width = w

    ws.cell(row=4 + len(info_list) + 1, column=1,
            value="Źródło: rcn.geoforum.pl | GUGiK | Dane bezpłatne od 13.02.2026"
            ).font = Font(name="Arial", size=8, italic=True, color="808080")


# ─── MAIN ─────────────────────────────────────────────────────────────────────

if __name__ == "__main__":
    # 1. Pobierz ZIP z GML-ami
    gml_pliki = pobierz_zip()

    if not gml_pliki:
        print("✗ Brak plików GML w archiwum.")
        exit(1)

    # 2. Parsuj każdy plik GML
    wszystkie_rekordy = []
    for nazwa, gml_bytes in gml_pliki.items():
        print(f"\nParsowanie: {nazwa}.gml ...")
        rekordy = parsuj_gml(gml_bytes)
        print(f"  → {len(rekordy)} rekordów")
        wszystkie_rekordy.extend(rekordy)

    if not wszystkie_rekordy:
        print("✗ Nie udało się wyekstrahować żadnych rekordów z GML.")
        print("  Możliwe przyczyny:")
        print("  - Niestandardowy namespace w GML powiatu bydgoskiego")
        print("  - Zagnieżdżenie danych głębiej niż oczekiwano")
        print("  Spróbuj otworzyć plik GML w QGIS lub przejrzeć surowy XML.")
        exit(1)

    df_all = pd.DataFrame(wszystkie_rekordy)
    print(f"\nWszystkie kolumny: {list(df_all.columns)}")
    print(f"Typy obiektów: {df_all['_typ'].value_counts().to_dict() if '_typ' in df_all.columns else 'n/d'}")

    # 3. Rozdziel na typy
    typ_col = "_typ" if "_typ" in df_all.columns else None

    def filtruj(pattern):
        if typ_col:
            mask = df_all[typ_col].str.contains(pattern, case=False, na=False)
            return przetworz_df(df_all[mask].copy(), pattern)
        return przetworz_df(df_all.copy(), pattern)

    df_lokale  = filtruj("Lokal|lokal")
    df_dzialki = filtruj("Dzialka|dzialka|grunt")
    df_budynki = filtruj("Budynek|budynek")

    # Fallback: jeśli nie ma rozdziału, wszystko w jednym arkuszu
    if len(df_lokale) == 0 and len(df_dzialki) == 0 and len(df_budynki) == 0:
        df_all_clean = przetworz_df(df_all.copy(), "all")
        arkusze = [("Wszystkie transakcje", df_all_clean)]
    else:
        arkusze = [
            ("Lokale", df_lokale),
            ("Dzialki", df_dzialki),
            ("Budynki", df_budynki),
        ]

    # 4. Zapis do Excela
    wb = Workbook()
    del wb["Sheet"]

    summary_info = []
    for ark_nazwa, df in arkusze:
        ws = wb.create_sheet(title=ark_nazwa)
        write_df_to_sheet(ws, df,
                          f"RCN | {ark_nazwa.upper()} | Powiat Bydgoski (0403)")
        if len(df) > 0:
            cena_med = int(df["cena_brutto_PLN"].median()) \
                if "cena_brutto_PLN" in df.columns else "–"
            m2_med   = int(df["cena_za_m2_PLN"].median()) \
                if "cena_za_m2_PLN" in df.columns else "–"
            z_cena   = int(df["cena_brutto_PLN"].notna().sum()) \
                if "cena_brutto_PLN" in df.columns else 0
            summary_info.append([ark_nazwa, len(df), z_cena, cena_med, m2_med])

    ws_sum = wb.create_sheet(title="Podsumowanie", index=0)
    write_summary(ws_sum, summary_info)

    wb.save(OUTPUT_PATH)
    print(f"\n✓ Zapisano: {OUTPUT_PATH}")
    for row in summary_info:
        print(f"  {row[0]:25s} {row[1]:6d} rekordów | z ceną: {row[2]:5d} | "
              f"mediana: {row[3]:>12} PLN | /m²: {row[4]:>8} PLN")