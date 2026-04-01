#!/usr/bin/env python3
# -*- coding: utf-8 -*-

from pathlib import Path
import os
import unicodedata

import tkinter as tk
from tkinter import filedialog, messagebox
import tkinter.ttk as ttk

import pandas as pd
import numpy as np
import manual
import subprocess
import sys
import threading  # <-- do wątku dla Automatu / Czyszczenia

from openpyxl import load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows

APP_TITLE = "PriceBot"

# --- nazwy arkuszy ---
RAPORT_SHEET = "raport"
RAPORT_ODF = "raport_odfiltrowane"


# ---------- Helpers nazewnicze ----------

def _norm(s: str) -> str:
    return (s or "").strip().lower().replace(" ", "").replace("\xa0", "").replace("\t", "")


def _plain(s: str) -> str:
    s = (s or "").lower()
    s = "".join(c for c in unicodedata.normalize("NFKD", s) if not unicodedata.combining(c))
    return s


def _find_col(cols, candidates):
    """Zwróć istniejącą kolumnę dopasowaną do listy kandydatów (po normalizacji / zawieraniu)."""
    norm_map = {_norm(c): c for c in cols}
    # dokładne
    for cand in candidates:
        key = _norm(cand)
        if key in norm_map:
            return norm_map[key]
    # "zawiera"
    for c in cols:
        if any(_norm(x) in _norm(c) for x in candidates):
            return c
    return None


def _trim_after_semicolon(val):
    if pd.isna(val):
        return ""
    s = str(val)
    if ";" in s:
        s = s.split(";", 1)[0].strip()
    return s


def _to_float_maybe(x):
    """Parsuje liczby typu '101,62 m²', '52 m2', '11 999 zł/m²' itd."""
    if pd.isna(x):
        return None
    s = str(x)

    # usuń jednostki
    for unit in ["m²", "m2", "zł/m²", "zł/m2", "zł"]:
        s = s.replace(unit, "")

    s = s.replace(" ", "").replace("\xa0", "")
    s = s.replace(",", ".")
    s = "".join(ch for ch in s if (ch.isdigit() or ch == "." or ch == "-"))
    try:
        return float(s) if s else None
    except Exception:
        return None


# ---------- Excel: czytaj/zapisuj TYLKO arkusz "raport" (bez kasowania innych) ----------

def _xlsx_has_sheet(path: Path, sheet_name: str) -> bool:
    try:
        wb = load_workbook(path, read_only=True, keep_vba=(path.suffix.lower() == ".xlsm"))
        return sheet_name in wb.sheetnames
    except Exception:
        return False


def _read_report_excel(path: Path, sheet_name: str = RAPORT_SHEET) -> pd.DataFrame:
    """Czyta WYŁĄCZNIE arkusz 'raport'. Jeśli nie istnieje – rzuca wyjątek."""
    if not _xlsx_has_sheet(path, sheet_name):
        raise RuntimeError(f"Plik nie zawiera arkusza '{sheet_name}'.")
    return pd.read_excel(path, sheet_name=sheet_name)


def _get_header_from_ws(ws) -> list[str]:
    header = []
    for cell in ws[1]:
        header.append(str(cell.value).strip() if cell.value is not None else "")
    while header and header[-1] == "":
        header.pop()
    return header


def ensure_raport_odfiltrowane(path: Path) -> None:
    """
    Gwarantuje istnienie arkusza 'raport_odfiltrowane' z SAMYMI nagłówkami,
    skopiowanymi z arkusza 'raport'. Nie rusza innych arkuszy.
    """
    if path.suffix.lower() not in (".xlsx", ".xlsm"):
        return

    keep_vba = path.suffix.lower() == ".xlsm"
    wb = load_workbook(path, keep_vba=keep_vba)

    # 1) zapewnij raport
    if RAPORT_SHEET not in wb.sheetnames:
        if wb.sheetnames:
            wb[wb.sheetnames[0]].title = RAPORT_SHEET
        else:
            wb.create_sheet(RAPORT_SHEET)

    ws_r = wb[RAPORT_SHEET]
    header = _get_header_from_ws(ws_r)

    # 2) zapewnij raport_odfiltrowane
    if RAPORT_ODF in wb.sheetnames:
        ws_o = wb[RAPORT_ODF]
    else:
        ws_o = wb.create_sheet(RAPORT_ODF)

    ws_o.sheet_state = "visible"

    # 3) wyczyść wszystko w arkuszu i wpisz tylko nagłówek
    if ws_o.max_row >= 1:
        ws_o.delete_rows(1, ws_o.max_row)

    # wpisz nagłówek
    for c, name in enumerate(header, start=1):
        ws_o.cell(row=1, column=c).value = name

    wb.save(path)


def _write_df_to_sheet_preserve(path: Path, df: pd.DataFrame, sheet_name: str = RAPORT_SHEET) -> None:
    """
    Zapisuje DataFrame do jednego arkusza (sheet_name) w pliku XLSX/XLSM
    i NIE dotyka pozostałych arkuszy (np. 'raport_odfiltrowane').
    Dodatkowo dopilnowuje, żeby 'raport_odfiltrowane' istniał i miał nagłówki.
    """
    keep_vba = path.suffix.lower() == ".xlsm"
    wb = load_workbook(path, keep_vba=keep_vba)

    if sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        ws.delete_rows(1, ws.max_row)
    else:
        ws = wb.create_sheet(sheet_name)

    for r in dataframe_to_rows(df, index=False, header=True):
        ws.append(r)

    wb.save(path)

    # dopilnuj raport_odfiltrowane
    try:
        ensure_raport_odfiltrowane(path)
    except Exception:
        pass


# ---------- USTAWIENIA PODGLĄDU ----------

PREVIEW_SPEC = [
    ("Nr KW", ["Nr KW", "nr_ksiegi", "nrksiegi", "nr księgi", "numer księgi"]),
    ("Województwo", ["Województwo", "wojewodztwo", "woj"]),
    ("Powiat", ["Powiat"]),
    ("Gmina", ["Gmina"]),
    ("Miejscowość", ["Miejscowość", "Miejscowosc", "Miasto"]),
    ("Dzielnica", ["Dzielnica", "Osiedle"]),
    ("Ulica", ["Ulica", "Ulica(dla budynku)", "Ulica(dla lokalu)"]),
    ("Obszar", [
        "Obszar", "metry", "powierzchnia",
        "Nr działek po średniku",
        "Nr działek", "Obręb po średniku", "Obręb"
    ]),
]

HIDDEN_PREVIEW_COLS = {_norm("Typ Księgi"), _norm("Stan Księgi")}

VALUE_COLS = [
    "Średnia cena za m2 ( z bazy)",
    "Średnia skorygowana cena za m2",
    "Statystyczna wartość nieruchomości",
]

# mapa nazw filtrów → skrypt
FILTER_SCRIPTS = {
    "Brak filtra": None,
    "Jeden właściciel": ["jeden_właściciel.py", "jeden_wlasciciel.py"],
    "LOKAL MIESZKALNY": ["LOKAL_MIESZKALNY.py", "lokal_mieszkalny.py"],
    "Jeden właściciel + LOKAL MIESZKALNY": [
        "jeden_właściciel_i_LOKAL_MIESZKALNY.py",
        "jeden_wlasciciel_i_lokal_mieszkalny.py",
    ],
    "Cofnij filtr": ["cofnij.py"],
}


# ---------- Główna klasa ----------

class App(tk.Tk):
    def __init__(self) -> None:
        super().__init__()
        self.title(APP_TITLE)
        self.minsize(980, 460)

        self.df: pd.DataFrame | None = None
        self.input_path: Path | None = None
        self.current_idx: int | None = None

        # Ustawienia
        self.input_file_var = tk.StringVar(value="")
        self.folder_var = tk.StringVar(value=str(Path.home()))  # baza: tu jest Polska.xlsx
        self.output_folder_var = tk.StringVar(value="")  # folder zapisu (Nr KW).xlsx
        self.margin_m2_var = tk.DoubleVar(value=15.0)  # okno ± m²
        self.margin_pct_var = tk.DoubleVar(value=15.0)  # obniżka % ceny
        self.filter_choice_var = tk.StringVar(value="Brak filtra")
        self.rows_count_var = tk.StringVar(value="Wiersze w raporcie: -")

        self.goto_kw_var = tk.StringVar(value="")
        self._kw_index: dict[str, int] = {}

        # --- UI ---
        root = ttk.Frame(self, padding=10)
        root.pack(fill="both", expand=True)

        # ---------- Plik wejściowy ----------
        group_in = ttk.LabelFrame(root, text="Plik raportu (wejście)")
        group_in.pack(fill="x")
        row_in = ttk.Frame(group_in)
        row_in.pack(fill="x", padx=8, pady=6)

        ttk.Entry(row_in, textvariable=self.input_file_var).pack(side="left", fill="x", expand=True)
        ttk.Button(row_in, text="Wybierz...", command=self.choose_input_file).pack(side="left", padx=(8, 0))

        # PRZYCISK CZYSZCZENIE PLIKU (kolory + logi)
        self.clean_btn = tk.Button(row_in, text="Czyszczenie Pliku", command=self.clean_input_file)
        self.clean_btn.pack(side="left", padx=(8, 0))

        # ---------- Folder bazowy ----------
        group_base = ttk.LabelFrame(root, text="Miejsce tworzenia plików i folderów")
        group_base.pack(fill="x", pady=(8, 0))
        row_base = ttk.Frame(group_base)
        row_base.pack(fill="x", padx=8, pady=6)
        ttk.Entry(row_base, textvariable=self.folder_var).pack(side="left", fill="x", expand=True)
        ttk.Button(row_base, text="Wybierz folder...", command=self.choose_base_folder).pack(side="left", padx=(8, 0))
        ttk.Button(row_base, text="Przygotowanie Aplikacji", command=self.prepare_app).pack(side="left", padx=(8, 0))

        # ---------- Baza danych ----------
        group_db = ttk.LabelFrame(root, text="Baza danych")
        group_db.pack(fill="x", pady=(8, 0))
        row_db = ttk.Frame(group_db)
        row_db.pack(fill="x", padx=8, pady=6)
        ttk.Button(row_db, text="Baza danych", command=self.run_bazadanych).pack(side="left")

        # ---------- Filtry ----------
        group_flt = ttk.LabelFrame(root, text="Filtry (opcjonalne)")
        group_flt.pack(fill="x", pady=(8, 0))
        row_flt = ttk.Frame(group_flt)
        row_flt.pack(fill="x", padx=8, pady=6)

        ttk.Label(row_flt, text="Wybierz filtr:").pack(side="left")
        cmb = ttk.Combobox(
            row_flt,
            textvariable=self.filter_choice_var,
            values=list(FILTER_SCRIPTS.keys()),
            state="readonly",
            width=35
        )
        cmb.pack(side="left", padx=(6, 6))
        cmb.current(0)
        ttk.Button(row_flt, text="Użyj filtru", command=self.apply_filter).pack(side="left")
        ttk.Button(row_flt, text="Odśwież", command=self.refresh_preview).pack(side="left", padx=(6, 0))
        ttk.Label(row_flt, textvariable=self.rows_count_var).pack(side="left", padx=(12, 0))

        # ---------- Folder wyników ----------
        group_out = ttk.LabelFrame(root, text="Folder zapisu wyników")
        group_out.pack(fill="x", pady=(8, 0))
        row_out = ttk.Frame(group_out)
        row_out.pack(fill="x", padx=8, pady=6)
        ttk.Entry(row_out, textvariable=self.output_folder_var).pack(side="left", fill="x", expand=True)
        ttk.Button(row_out, text="Wybierz folder...", command=self.choose_output_folder).pack(side="left", padx=(8, 0))

        # ---------- Sterowanie ----------
        group_nav = ttk.LabelFrame(root, text="Sterowanie")
        group_nav.pack(fill="x", pady=(8, 0))
        row_nav = ttk.Frame(group_nav)
        row_nav.pack(fill="x", padx=8, pady=6)

        ttk.Button(row_nav, text="‹ Poprzedni", command=self.prev_row).pack(side="left")
        ttk.Button(row_nav, text="Następny ›", command=self.next_row).pack(side="left", padx=(6, 0))

        ttk.Separator(row_nav, orient="vertical").pack(side="left", fill="y", padx=10)

        ttk.Label(row_nav, text="Przejdź do Nr KW:").pack(side="left")
        kw_entry = ttk.Entry(row_nav, textvariable=self.goto_kw_var, width=30)
        kw_entry.pack(side="left", padx=(6, 6))
        kw_entry.bind("<Return>", lambda _e: self.goto_kw())
        ttk.Button(row_nav, text="Idź", command=self.goto_kw).pack(side="left")

        # ---------- Obliczenia ----------
        group_calc = ttk.LabelFrame(root, text="Obliczenia")
        group_calc.pack(fill="x", pady=(8, 0))
        row_calc = ttk.Frame(group_calc)
        row_calc.pack(fill="x", padx=8, pady=6)

        ttk.Label(row_calc, text="Pomiar brzegowy metrażu (± m²):").pack(side="left")
        ttk.Spinbox(
            row_calc,
            from_=0.0, to=200.0, increment=0.5,
            width=6, textvariable=self.margin_m2_var
        ).pack(side="left", padx=(6, 14))

        ttk.Label(row_calc, text="Obniżka ceny (%):").pack(side="left")
        ttk.Spinbox(
            row_calc,
            from_=0.0, to=100.0, increment=0.5,
            width=6, textvariable=self.margin_pct_var
        ).pack(side="left", padx=(6, 14))

        ttk.Button(
            row_calc,
            text="Oblicz i zapisz ten wiersz",
            command=self.calc_and_save_row
        ).pack(side="left")

        self.automat_btn = tk.Button(row_calc, text="Automat", command=self.automate)
        self.automat_btn.pack(side="left", padx=(6, 0))

        # ---------- Podgląd ----------
        group_preview = ttk.LabelFrame(root, text="Bieżący wiersz (podgląd)")
        group_preview.pack(fill="both", expand=True, pady=(8, 0))
        self.preview_label = ttk.Label(
            group_preview,
            text="{Wybierz plik raportu}",
            anchor="w",
            justify="left"
        )
        self.preview_label.pack(fill="both", expand=True, padx=8, pady=6)

    # ---------- uruchamianie zewnętrznych skryptów ----------

    def _run_script(self, candidates: list[str], extra_args: list[str] | None = None):
        if not candidates:
            return
        extra_args = extra_args or []
        here = Path(__file__).resolve().parent
        for name in candidates:
            script = here / name
            if script.exists():
                try:
                    subprocess.Popen(
                        [sys.executable, str(script), *extra_args],
                        cwd=str(here),
                        close_fds=(os.name != "nt"),
                        creationflags=(subprocess.CREATE_NO_WINDOW if os.name == "nt" else 0),
                    )
                    messagebox.showinfo("Uruchomiono", f"Start: {script.name}")
                    return
                except Exception as e:
                    messagebox.showerror("Błąd uruchamiania", f"Nie udało się uruchomić {script.name}:\n{e}")
                    return
        messagebox.showerror("Brak pliku", f"Nie znaleziono żadnego ze skryptów: {', '.join(candidates)}")

    # ---------- GUI actions ----------

    def choose_input_file(self):
        path = filedialog.askopenfilename(
            title="Wybierz plik raportu (CSV/XLSX/XLSM)",
            filetypes=[
                ("CSV", "*.csv"),
                ("Excel", "*.xlsx;*.xlsm"),
                ("Wszystkie pliki", "*.*"),
            ],
        )
        if not path:
            return
        self.input_file_var.set(path)
        self.input_path = Path(path)
        self.load_dataframe(self.input_path)
        self.goto_row(0)

    def load_dataframe(self, path: Path):
        try:
            if path.suffix.lower() in (".xlsx", ".xlsm"):
                # ⛔ Podgląd ma być TYLKO arkusza 'raport'
                self.df = _read_report_excel(path, sheet_name=RAPORT_SHEET)

                self.update_rows_count()

                # ✅ dopilnuj arkusza raport_odfiltrowane (techniczne)
                try:
                    ensure_raport_odfiltrowane(path)
                except Exception:
                    pass
            else:
                # CSV nie ma arkuszy — podgląd OK
                self.df = pd.read_csv(path, sep=None, engine="python")

            self.rebuild_kw_index()
            self.update_rows_count()
        except Exception as e:
            messagebox.showerror(
                "Błąd odczytu",
                f"Nie mogę wczytać arkusza '{RAPORT_SHEET}' z pliku:{path}{e}"
            )
            self.df = None
            self.current_idx = None
            self.preview_label.config(text="{Brak arkusza 'raport'}")

        self.update_rows_count()

    def update_rows_count(self):
        """Aktualizuje licznik wierszy w arkuszu 'raport' (do wyliczenia)."""
        if getattr(self, "df", None) is None:
            self.rows_count_var.set("Wiersze w raporcie: -")
        else:
            self.rows_count_var.set(f"Wiersze w raporcie: {len(self.df)}")

    def refresh_preview(self):
        """Odświeża podgląd po użyciu filtru: ponownie wczytuje arkusz 'raport' i aktualizuje licznik."""
        in_path = self.input_file_var.get().strip()
        if not in_path:
            messagebox.showerror("Odśwież", "Najpierw wybierz plik raportu (u góry).")
            return

        self.input_path = Path(in_path)
        self.load_dataframe(self.input_path)

        if self.df is None or len(self.df.index) == 0:
            self.current_idx = None
            self.preview_label.config(text="{Brak danych w arkuszu 'raport'}")
            messagebox.showinfo("Odśwież", "Odświeżono podgląd. Arkusz 'raport' jest pusty.")
            return

        i = 0 if self.current_idx is None else min(self.current_idx, len(self.df.index) - 1)
        self.goto_row(i)
        messagebox.showinfo("Odśwież", f"Odświeżono podgląd. Wierszy w arkuszu '{RAPORT_SHEET}': {len(self.df)}")

    # ---------- CZYSZCZENIE PLIKU Z LOGIEM ----------

    def clean_input_file(self):
        in_path = self.input_file_var.get().strip()
        if not in_path:
            messagebox.showerror("Czyszczenie", "Najpierw wybierz plik raportu (u góry).")
            return

        here = Path(__file__).resolve().parent
        candidates = ["CzyszczenieAdresu.py", "czyszczeniadresu.py"]

        script_path = None
        for name in candidates:
            p = here / name
            if p.exists():
                script_path = p
                break

        if script_path is None:
            messagebox.showerror("Czyszczenie", f"Nie znaleziono żadnego ze skryptów: {', '.join(candidates)}")
            return

        try:
            self.clean_btn.config(bg="#f7e26b", activebackground="#f5d742")
        except Exception:
            pass

        def worker():
            try:
                env = os.environ.copy()
                # Wymuś UTF-8 w procesie potomnym, żeby GUI nie wywalało się na polskich znakach
                env["PYTHONIOENCODING"] = "utf-8"

                proc = subprocess.Popen(
                    [sys.executable, str(script_path), in_path],
                    cwd=str(here),
                    stdout=subprocess.PIPE,
                    stderr=subprocess.PIPE,
                    text=True,
                    encoding="utf-8",
                    errors="replace",
                    env=env,
                    close_fds=(os.name != "nt"),
                    creationflags=(subprocess.CREATE_NO_WINDOW if os.name == "nt" else 0),
                )
                out, err = proc.communicate()
                rc = proc.returncode
            except Exception as e:
                err_msg = str(e)

                def on_error(msg=err_msg):
                    try:
                        self.clean_btn.config(bg="#f28b82", activebackground="#ea4335")
                    except Exception:
                        pass
                    messagebox.showerror("Czyszczenie", f"Nie udało się uruchomić {script_path.name}:\n{msg}")

                self.after(0, on_error)
                return

            def on_done():
                if rc == 0:
                    try:
                        self.clean_btn.config(bg="#8ef98e", activebackground="#76e476")
                    except Exception:
                        pass
                    msg = f"Zakończono działanie {script_path.name}.\nPrzetworzony plik:\n{in_path}"
                    log = (out or "").strip()
                    if log:
                        msg += "\n\nLOG:\n" + log[-800:]
                    messagebox.showinfo("Czyszczenie", msg)
                else:
                    try:
                        self.clean_btn.config(bg="#f28b82", activebackground="#ea4335")
                    except Exception:
                        pass
                    log = ((err or "") + "\n" + (out or "")).strip() or "(brak tekstu błędu na stdout/stderr)"
                    if len(log) > 1500:
                        log = "...[ucięto początek]\n" + log[-1500:]
                    messagebox.showerror("Czyszczenie", f"{script_path.name} zakończył się błędem (kod {rc}).\n\n{log}")

            self.after(0, on_done)

        threading.Thread(target=worker, daemon=True).start()

    def choose_base_folder(self):
        d = filedialog.askdirectory(title="Wybierz folder bazowy", initialdir=self.folder_var.get())
        if d:
            self.folder_var.set(d)

    def prepare_app(self):
        base = Path(self.folder_var.get()).resolve()
        for p in ["linki", "województwa", "logs"]:
            (base / p).mkdir(parents=True, exist_ok=True)

        # ✅ dopilnuj raport_odfiltrowane zawsze
        if self.input_path and self.input_path.suffix.lower() in (".xlsx", ".xlsm"):
            try:
                ensure_raport_odfiltrowane(self.input_path)
            except Exception:
                pass

        if self.input_file_var.get().strip():
            self.add_value_columns_to_input()

        messagebox.showinfo("Przygotowanie Aplikacji", f"Przygotowano strukturę w:\n{base}")

    def choose_output_folder(self):
        d = filedialog.askdirectory(
            title="Wybierz folder zapisu wyników",
            initialdir=self.output_folder_var.get() or self.folder_var.get(),
        )
        if d:
            self.output_folder_var.set(d)

    # ✅ POPRAWIONE: dodaje kolumny przez openpyxl do arkusza 'raport' bez kasowania innych arkuszy
    def add_value_columns_to_input(self):
        """
        Dodaje 3 kolumny wartości do arkusza 'raport' w pliku raportowym,
        nie kasując innych arkuszy (np. 'raport_odfiltrowane').
        Wstawia je zaraz za 'Czy udziały?' jeśli istnieje.
        Dodatkowo dopilnowuje istnienia arkusza 'raport_odfiltrowane'.
        """
        in_path_str = self.input_file_var.get().strip()
        if not in_path_str:
            messagebox.showerror("Kolumny", "Najpierw wybierz plik raportu (u góry).")
            return

        path = Path(in_path_str)
        if not path.exists():
            messagebox.showerror("Kolumny", f"Plik raportu nie istnieje:\n{path}")
            return

        if path.suffix.lower() not in (".xlsx", ".xlsm"):
            messagebox.showerror("Kolumny", "Ta funkcja działa dla plików Excel (.xlsx/.xlsm).")
            return

        try:
            keep_vba = path.suffix.lower() == ".xlsm"
            wb = load_workbook(path, keep_vba=keep_vba)

            # wybierz/utwórz arkusz 'raport'
            if RAPORT_SHEET in wb.sheetnames:
                ws = wb[RAPORT_SHEET]
            else:
                ws = wb[wb.sheetnames[0]]
                if RAPORT_SHEET in wb.sheetnames:
                    wb.remove(wb[RAPORT_SHEET])
                ws.title = RAPORT_SHEET

            # odczytaj nagłówek
            header = []
            for cell in ws[1]:
                header.append(str(cell.value).strip() if cell.value is not None else "")
            while header and header[-1] == "":
                header.pop()

            # pozycja insertu
            col_udzialy = None
            for i, name in enumerate(header):
                if _norm(name) in (_norm("Czy udziały?"), _norm("Czy udzialy")):
                    col_udzialy = i
                    break
            insert_pos = (col_udzialy + 1) if col_udzialy is not None else len(header)

            to_add = [c for c in VALUE_COLS if c not in header]
            if not to_add:
                wb.save(path)
                try:
                    ensure_raport_odfiltrowane(path)
                except Exception:
                    pass
                messagebox.showinfo("Kolumny", "Kolumny wartości już istnieją w arkuszu 'raport'.")
                return

            for k, col_name in enumerate(to_add):
                ws.insert_cols(insert_pos + 1 + k)
                ws.cell(row=1, column=insert_pos + 1 + k).value = col_name

            wb.save(path)

            # dopilnuj odfiltrowane
            try:
                ensure_raport_odfiltrowane(path)
            except Exception:
                pass

        except PermissionError:
            messagebox.showerror(
                "Kolumny",
                "Nie udało się dodać kolumn — plik jest pewnie otwarty w Excelu.\n"
                "Zamknij plik i spróbuj ponownie.",
            )
            return
        except Exception as e:
            messagebox.showerror("Kolumny", f"Nie udało się dodać kolumn:\n{e}")
            return

        messagebox.showinfo(
            "Kolumny",
            "Dodano brakujące kolumny wartości do arkusza 'raport' bez kasowania innych arkuszy.\n"
            "Dodatkowo przygotowano arkusz 'raport_odfiltrowane' (same nagłówki).",
        )

    def run_bazadanych(self):
        root_dir = Path(self.folder_var.get()).resolve()
        if not root_dir.exists():
            messagebox.showerror("Baza danych", f"Folder bazowy nie istnieje:\n{root_dir}")
            return
        try:
            from bazadanych import open_ui
        except Exception as e:
            messagebox.showerror("Baza danych", f"Nie mogę zaimportować bazadanych.py:\n{e}")
            return
        open_ui(root_dir, parent=self)

    # --------- AUTOMAT ----------

    def automate(self):
        raport = self.input_file_var.get().strip()
        baza = self.folder_var.get().strip()

        if not raport:
            messagebox.showerror("Automat", "Najpierw wybierz plik raportu (u góry).")
            return
        if not baza:
            messagebox.showerror("Automat", "Najpierw ustaw folder bazowy (tam gdzie Polska.xlsx).")
            return

        try:
            self.automat_btn.config(bg="#f7e26b", activebackground="#f5d742")
        except Exception:
            pass

        try:
            import automat
        except Exception as e:
            try:
                self.automat_btn.config(bg="", activebackground="")
            except Exception:
                pass
            messagebox.showerror("Automat", f"Nie mogę zaimportować automat.py:\n{e}")
            return

        def worker():
            try:
                rc = automat.main(["automat.py", raport, baza])
            except Exception as e:
                err_msg = str(e)

                def on_error(msg=err_msg):
                    try:
                        self.automat_btn.config(bg="", activebackground="")
                    except Exception:
                        pass
                    messagebox.showerror("Automat", f"Błąd działania automat.py:\n{msg}")

                self.after(0, on_error)
                return

            def on_done():
                if rc == 0:
                    try:
                        self.automat_btn.config(bg="#8ef98e", activebackground="#76e476")
                    except Exception:
                        pass
                    messagebox.showinfo(
                        "Automat",
                        "Zakończono działanie automat.py.\nWyniki powinny być wpisane do kolumn w raporcie.",
                    )
                else:
                    try:
                        self.automat_btn.config(bg="#f28b82", activebackground="#ea4335")
                    except Exception:
                        pass
                    messagebox.showerror("Automat", "automat.py zakończył się błędem (kod != 0). Sprawdź logi.")

            self.after(0, on_done)

        threading.Thread(target=worker, daemon=True).start()

    def apply_filter(self):
        choice = self.filter_choice_var.get()
        scripts = FILTER_SCRIPTS.get(choice)
        if not scripts:
            messagebox.showinfo("Filtry", "Wybrano 'Brak filtra' – nic nie uruchamiam.")
            return
        in_path = self.input_file_var.get().strip()
        if not in_path:
            messagebox.showerror("Filtry", "Najpierw wybierz plik raportu (u góry).")
            return
        self._run_script(scripts, extra_args=["--in", in_path])

    # ---------- Nr KW: indeks + skok ----------

    def _kw_key(self, v) -> str:
        """Normalizacja wartości Nr KW do wyszukiwania (bez spacji, uppercase, obcięcie po ';')."""
        s = _trim_after_semicolon(v)
        s = str(s).strip() if s is not None else ""
        if not s:
            return ""
        return s.replace(" ", "").replace("\xa0", "").upper()

    def rebuild_kw_index(self) -> None:
        """Buduje mapę Nr KW -> indeks wiersza (pierwsze wystąpienie)."""
        self._kw_index = {}
        if self.df is None or len(self.df.index) == 0:
            return
        col = _find_col(
            self.df.columns,
            ["Nr KW", "nr kw", "NrKW", "nr_ksiegi", "nrksiegi", "nr księgi", "numer księgi"]
        )
        if not col:
            return
        for i, v in enumerate(self.df[col].tolist()):
            key = self._kw_key(v)
            if key and key not in self._kw_index:
                self._kw_index[key] = i

    def goto_kw(self) -> None:
        """Przejdź do wiersza o podanym Nr KW."""
        if self.df is None:
            messagebox.showinfo("Nawigacja", "Najpierw wybierz plik raportu.")
            return

        target_raw = self.goto_kw_var.get()
        target = self._kw_key(target_raw)

        if not target:
            messagebox.showinfo("Nawigacja", "Wpisz Nr KW, żeby przejść do konkretnego wiersza.")
            return

        # przebuduj indeks, jeśli pusty (np. po odświeżeniu)
        if not getattr(self, "_kw_index", None):
            self.rebuild_kw_index()

        idx = self._kw_index.get(target)
        if idx is None:
            # awaryjnie: liniowe wyszukiwanie (gdyby kolumna się inaczej nazywała)
            col = _find_col(
                self.df.columns,
                ["Nr KW", "nr kw", "NrKW", "nr_ksiegi", "nrksiegi", "nr księgi", "numer księgi"]
            )
            if col:
                for i, v in enumerate(self.df[col].tolist()):
                    if self._kw_key(v) == target:
                        idx = i
                        break

        if idx is None:
            messagebox.showinfo("Nawigacja", f"Nie znaleziono Nr KW: {target_raw.strip()}")
            return

        self.goto_row(int(idx))

    # ---------- Nawigacja ----------

    def prev_row(self):
        if self.df is None:
            messagebox.showinfo("Nawigacja", "Najpierw wybierz plik raportu.")
            return
        i = 0 if self.current_idx is None else max(0, self.current_idx - 1)
        self.goto_row(i)

    def next_row(self):
        if self.df is None:
            messagebox.showinfo("Nawigacja", "Najpierw wybierz plik raportu.")
            return
        n = len(self.df.index)
        i = 0 if self.current_idx is None else min(n - 1, self.current_idx + 1)
        self.goto_row(i)

    # ---------- PODGLĄD ----------

    def goto_row(self, i: int):
        if self.df is None or i < 0 or i >= len(self.df.index):
            return
        self.current_idx = i
        row = self.df.iloc[i]
        lines = [f"Wiersz {i + 1}/{len(self.df)}"]
        for label, candidates in PREVIEW_SPEC:
            if any(_norm(c) in HIDDEN_PREVIEW_COLS for c in candidates):
                continue
            col = _find_col(self.df.columns, candidates)
            val = _trim_after_semicolon(row[col]) if col else ""
            lines.append(f"• {label}: {val}")
        self.preview_label.config(text="\n".join(lines))

    # ---------- KALKULACJA + ZAPIS ----------

    def calc_and_save_row(self):
        """
        Ręczne liczenie (NOWY algorytm) — logika jest w manual.py.
        Przycisk w GUI: "Oblicz i zapisz ten wiersz".
        """
        if self.df is None or self.current_idx is None:
            messagebox.showinfo("Zapis", "Najpierw wybierz plik raportu i wiersz.")
            return
        if not self.output_folder_var.get() and not self.folder_var.get():
            messagebox.showerror("Brak folderu", "Wybierz 'Folder zapisu wyników'.")
            return

        base_dir = Path(self.folder_var.get()).resolve()
        out_dir = Path(self.output_folder_var.get() or self.folder_var.get()).resolve()

        try:
            res = manual.compute_and_save_row(
                df_report=self.df,
                idx=int(self.current_idx),
                base_dir=base_dir,
                out_dir=out_dir,
                margin_m2_default=float(self.margin_m2_var.get() or 15.0),
                margin_pct_default=float(self.margin_pct_var.get() or 15.0),
                min_hits=5,
            )
        except manual.ManualUserError as e:
            messagebox.showerror("Błąd", str(e))
            return
        except Exception as e:
            messagebox.showerror("Błąd", f"Nie udało się przeliczyć wiersza:\n{e}")
            return

        # ✅ Zapis raportu: XLSX tylko arkusz 'raport' (bez kasowania innych arkuszy)
        try:
            if self.input_path and self.input_path.suffix.lower() in (".xlsx", ".xlsm"):
                _write_df_to_sheet_preserve(self.input_path, self.df, sheet_name=RAPORT_SHEET)
            elif self.input_path and self.input_path.suffix.lower() == ".csv":
                self.df.to_csv(self.input_path, index=False, encoding="utf-8-sig")
        except Exception as e:
            messagebox.showwarning(
                "Zapis raportu",
                f"Wyliczono wartości, ale nie udało się zapisać raportu:\n{self.input_path}\n\n{e}",
            )

        # komunikat
        msg = []
        if res.get("out_path"):
            msg.append(f"Zapisano dobrane rekordy do: {res['out_path']}")
        if isinstance(res.get("avg"), (int, float)):
            msg.append("Średnia cena/m²: " + f"{res['avg']:,.2f}".replace(",", " ").replace(".", ","))
        if isinstance(res.get("corrected"), (int, float)) and isinstance(res.get("avg"), (int, float)) and res.get(
                "corrected") != res.get("avg"):
            pct = None
            margins = res.get("margins")
            if isinstance(margins, (tuple, list)) and len(margins) >= 2:
                pct = float(margins[1])
            if pct is None:
                try:
                    pct = float(self.margin_pct_var.get() or 0.0)
                except Exception:
                    pct = 0.0
            msg.append(
                f"Średnia po obniżce ({pct:.1f}%): " + f"{res['corrected']:,.2f}".replace(",", " ").replace(".", ","))
        if isinstance(res.get("value"), (int, float)):
            msg.append("Statystyczna wartość: " + f"{res['value']:,.2f}".replace(",", " ").replace(".", ","))
        if res.get("stage"):
            msg.append(f"Etap doboru: {res['stage']} (trafień: {int(res.get('hits', 0))})")

        messagebox.showinfo("Zakończono", "\n".join(msg) if msg else "Zakończono.")


def main():
    app = App()
    app.mainloop()


if __name__ == "__main__":
    main()
