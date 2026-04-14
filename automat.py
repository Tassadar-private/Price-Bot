#!/usr/bin/env python3
# -*- coding: utf-8 -*-
from __future__ import annotations
import logging
from pathlib import Path
import sys
import pandas as pd
import numpy as np
import tkinter as tk
from tkinter import ttk
from openpyxl import load_workbook
import automat1 as a1
from utils import setup_logging

logger = logging.getLogger(__name__)
def _pick_report_sheet_name(xlsx_path: Path, preferred: str = "raport") -> str:
    wb = load_workbook(xlsx_path)
    if preferred in wb.sheetnames:
        return preferred
    return wb.sheetnames[0] if wb.sheetnames else preferred

def save_report_sheet_only(xlsx_path: Path, df_report: pd.DataFrame, sheet_name: str = "raport") -> None:
    wb = load_workbook(xlsx_path)

    if sheet_name not in wb.sheetnames:
        if wb.sheetnames:
            sheet_name = wb.sheetnames[0]
        else:
            wb.create_sheet(sheet_name)

    ws = wb[sheet_name]

    if ws.max_row >= 1:
        ws.delete_rows(1, ws.max_row)

    for c, name in enumerate(df_report.columns.tolist(), start=1):
        ws.cell(row=1, column=c, value=name)

    for r_idx, row in enumerate(df_report.values.tolist(), start=2):
        for c_idx, val in enumerate(row, start=1):
            ws.cell(row=r_idx, column=c_idx, value=val)

    wb.save(xlsx_path)

def configure_margins_gui():
    root = tk.Tk()
    root.title("Ustawienia progów ludności")
    root.resizable(False, False)

    ttk.Label(root, text="Minimalna ludność").grid(row=0, column=0, padx=4, pady=4)
    ttk.Label(root, text="Maksymalna ludność").grid(row=0, column=1, padx=4, pady=4)
    ttk.Label(root, text="Pomiar brzegowy m²").grid(row=0, column=2, padx=4, pady=4)
    ttk.Label(root, text="% negocjacyjny").grid(row=0, column=3, padx=4, pady=4)

    entries_m2: list[ttk.Entry] = []
    entries_pct: list[ttk.Entry] = []

    def _fmt_pop(x):
        if x is None:
            return "∞"
        try:
            x = int(x)
        except (ValueError, TypeError):
            return str(x)
        return f"{x:,}".replace(",", " ")

    for i, (low, high, m2, pct) in enumerate(a1.POP_MARGIN_RULES, start=1):
        ttk.Label(root, text=_fmt_pop(low)).grid(row=i, column=0, padx=4, pady=2, sticky="e")
        ttk.Label(root, text=_fmt_pop(high)).grid(row=i, column=1, padx=4, pady=2, sticky="e")

        e_m2 = ttk.Entry(root, width=8, justify="right")
        e_m2.insert(0, str(m2))
        e_m2.grid(row=i, column=2, padx=4, pady=2)
        entries_m2.append(e_m2)

        e_pct = ttk.Entry(root, width=8, justify="right")
        e_pct.insert(0, str(pct))
        e_pct.grid(row=i, column=3, padx=4, pady=2)
        entries_pct.append(e_pct)

    result = {"ok": False, "rules": a1.POP_MARGIN_RULES}

    def on_ok():
        new_rules = []
        for (low, high, default_m2, default_pct), e_m2, e_pct in zip(a1.POP_MARGIN_RULES, entries_m2, entries_pct):
            raw_m2 = e_m2.get().strip().replace(" ", "").replace(",", ".")
            raw_pct = e_pct.get().strip().replace(" ", "").replace(",", ".")
            try:
                m2_val = float(raw_m2) if raw_m2 else float(default_m2)
            except (ValueError, TypeError):
                m2_val = float(default_m2)
            try:
                pct_val = float(raw_pct) if raw_pct else float(default_pct)
            except (ValueError, TypeError):
                pct_val = float(default_pct)
            new_rules.append((low, high, m2_val, pct_val))
        result["ok"] = True
        result["rules"] = new_rules
        root.destroy()

    def on_cancel():
        result["ok"] = False
        root.destroy()

    btn_frame = ttk.Frame(root)
    btn_frame.grid(row=len(a1.POP_MARGIN_RULES) + 1, column=0, columnspan=4, pady=(8, 8))
    ttk.Button(btn_frame, text="Anuluj", command=on_cancel).pack(side="right", padx=4)
    ttk.Button(btn_frame, text="Start", command=on_ok).pack(side="right", padx=4)

    root.update_idletasks()
    w, h = root.winfo_width(), root.winfo_height()
    x = (root.winfo_screenwidth() - w) // 2
    y = (root.winfo_screenheight() - h) // 2
    root.geometry(f"{w}x{h}+{x}+{y}")
    root.mainloop()

    return result["rules"] if result["ok"] else None

def main(argv=None) -> int:
    setup_logging()

    if argv is None:
        argv = sys.argv

    if len(argv) < 3:
        logger.error("Uzycie: automat.py RAPORT_PATH BAZA_FOLDER")
        return 1

    raport_path = Path(argv[1]).resolve()
    baza_folder = Path(argv[2]).resolve()

    if not raport_path.exists():
        logger.error("Nie znaleziono raportu: %s", raport_path)
        return 1

    polska_path = baza_folder / "Polska.xlsx"
    if not polska_path.exists():
        logger.error("Nie znaleziono Polska.xlsx w folderze: %s", baza_folder)
        return 1

    margin_m2_default = 15.0
    margin_pct_default = 15.0

    try:
        new_rules = configure_margins_gui()
    except (tk.TclError, RuntimeError) as e:
        logger.error("Blad GUI progow ludnosci: %s", e)
        new_rules = a1.POP_MARGIN_RULES

    if new_rules is None:
        logger.info("Przerwano (Anuluj w oknie progow ludnosci).")
        return 1
    a1.POP_MARGIN_RULES = new_rules

    try:
        if len(a1.POP_MARGIN_RULES) >= 3:
            margin_m2_default = float(a1.POP_MARGIN_RULES[2][2])
            margin_pct_default = float(a1.POP_MARGIN_RULES[2][3])
    except (ValueError, TypeError, IndexError) as e:
        logger.warning("Nie udalo sie odczytac progow z POP_MARGIN_RULES: %s", e)

    try:
        df_pl = pd.read_excel(polska_path)
    except Exception as e:
        logger.error("Nie moge wczytac Polska.xlsx: %s — %s", polska_path, e)
        return 1

    col_area_pl = a1._find_col(df_pl.columns, ["metry", "powierzchnia", "Obszar", "obszar"])
    col_price_pl = a1._find_col(df_pl.columns, ["cena_za_metr", "cena za metr", "cena_za_m2", "cena_za_metr2", "cena za m2"])
    if not col_area_pl or not col_price_pl:
        logger.error("Polska.xlsx nie zawiera wymaganych kolumn metrazu / ceny.")
        return 1

    # indeksy dla szybkich dopasowań (kanonizacja + mapy miejscowości)
    pl_index = a1.build_polska_index(df_pl, col_area_pl, col_price_pl)

    is_excel = raport_path.suffix.lower() in [".xlsx", ".xlsm", ".xls"]
    try:
        if is_excel:
            sheet_to_read = _pick_report_sheet_name(raport_path, preferred="raport")
            df_raport = pd.read_excel(raport_path, sheet_name=sheet_to_read)
        else:
            df_raport = pd.read_csv(raport_path, sep=None, engine="python")
    except Exception as e:
        logger.error("Nie moge wczytac raportu: %s — %s", raport_path, e)
        return 1


    # zapewnij kolumny diagnostyczne + wynikowe (hits/stage + wycena)
    try:
        if hasattr(a1, "ensure_report_columns"):
            a1.ensure_report_columns(df_raport)
    except Exception as e:
        logger.warning("Nie moge przygotowac kolumn (hits/stage): %s", e)

    local_ludnosc = a1._find_ludnosc_csv(baza_folder=baza_folder, raport_path=raport_path, polska_path=polska_path)
    api_cache = baza_folder / "population_cache.csv"

    logger.info("local ludnosc.csv -> %s", local_ludnosc if local_ludnosc else "(NIE ZNALEZIONO)")

    pop_resolver = a1.PopulationResolver(local_csv=local_ludnosc, api_cache_csv=api_cache, use_api=True)

    logger.info("Start — liczba wierszy w raporcie: %d", len(df_raport.index))

    for idx in range(len(df_raport.index)):
        try:
            a1._process_row(
                df_raport=df_raport,
                idx=idx,
                pl=pl_index,
                margin_m2_default=margin_m2_default,
                margin_pct_default=margin_pct_default,
                pop_resolver=pop_resolver,
            )
        except Exception as e:
            logger.error("Blad przy wierszu %d: %s", idx + 1, e)

    try:
        # uporządkuj kolumny: Czy udziały? -> hits -> stage -> Średnia cena za m2 ( z bazy)
        try:
            if hasattr(a1, "reorder_report_columns"):
                df_raport = a1.reorder_report_columns(df_raport)
        except Exception as e:
            logger.warning("Nie moge uporzadkowac kolumn (hits/stage): %s", e)

        if is_excel:
            save_report_sheet_only(raport_path, df_raport, sheet_name="raport")
        else:
            df_raport.to_csv(raport_path, index=False, encoding="utf-8-sig")
    except Exception as e:
        logger.error("Nie udalo sie zapisac raportu: %s — %s", raport_path, e)
        return 1

    logger.info("Zakonczono — zapisano zmiany w pliku: %s", raport_path)
    return 0

if __name__ == '__main__':
    raise SystemExit(main())
