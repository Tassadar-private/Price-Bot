#!/usr/bin/env python3
# -*- coding: utf-8 -*-

from pathlib import Path
from typing import List
import argparse
import sys
import tkinter as tk
from tkinter import filedialog, messagebox
from openpyxl import load_workbook

RAPORT_SHEET = "raport"
RAPORT_ODF = "raport_odfiltrowane"
SUPPORTED = {".xlsx", ".xlsm"}

REQ_COLS: List[str] = [
    "Nr KW","Typ Księgi","Stan Księgi","Województwo","Powiat","Gmina",
    "Miejscowość","Dzielnica","Położenie","Nr działek po średniku",
    "Obręb po średniku","Ulica","Sposób korzystania","Obszar",
    "Ulica(dla budynku)","przeznaczenie (dla budynku)","Ulica(dla lokalu)",
    "Nr budynku( dla lokalu)","Przeznaczenie (dla lokalu)",
    "Cały adres (dla lokalu)","Czy udziały?"
]

VALUE_COLS: List[str] = [
    "Średnia cena za m2 ( z bazy)",
    "Średnia skorygowana cena za m2",
    "Statystyczna wartość nieruchomości",
]


def _read_header(ws) -> List[str]:
    header = []
    for cell in ws[1]:
        header.append(str(cell.value).strip() if cell.value is not None else "")
    while header and header[-1] == "":
        header.pop()
    return header


def _write_header(ws, header: List[str]) -> None:
    for i, name in enumerate(header, start=1):
        ws.cell(row=1, column=i).value = name


def _ensure_headers_on_report(ws, target_cols: List[str]) -> List[str]:
    existing = _read_header(ws)
    existing_set = {c for c in existing if c}

    final_header = existing[:]
    for name in target_cols:
        if name not in existing_set:
            final_header.append(name)
            existing_set.add(name)

    _write_header(ws, final_header)
    return final_header


def _ensure_headers_only_sheet(wb, title: str, header: List[str]):
    """
    Stabilnie:
    - jeśli arkusz istnieje -> NIE usuwaj, tylko wyczyść dane (wiersze 2+)
    - jeśli nie istnieje -> utwórz
    - zawsze wpisz nagłówek w wierszu 1
    - zawsze ustaw visible
    """
    if title in wb.sheetnames:
        ws = wb[title]
    else:
        ws = wb.create_sheet(title)

    ws.sheet_state = "visible"

    # Wyczyść dane od wiersza 2 w dół
    if ws.max_row >= 2:
        ws.delete_rows(2, ws.max_row - 1)

    # Wyczyść ewentualne stare kolumny nagłówka (żeby nie zostały śmieci po prawej)
    if ws.max_column > len(header):
        for c in range(len(header) + 1, ws.max_column + 1):
            ws.cell(row=1, column=c).value = None

    _write_header(ws, header)
    return ws


def ensure_report_columns(xlsx_path: Path | str) -> None:
    xlsx_path = Path(xlsx_path).expanduser().resolve()

    if not xlsx_path.exists():
        raise FileNotFoundError(f"Nie znaleziono pliku: {xlsx_path}")

    suf = xlsx_path.suffix.lower()
    if suf not in SUPPORTED:
        raise ValueError(f"Obsługiwane tylko pliki: {SUPPORTED}")

    # keep_vba dla .xlsm (żeby nic nie znikało)
    wb = load_workbook(xlsx_path, keep_vba=(suf == ".xlsm"))

    if not wb.sheetnames:
        raise ValueError("Brak arkuszy w pliku.")

    # wybór bazowego arkusza do zmiany nazwy na 'raport'
    if RAPORT_SHEET in wb.sheetnames:
        ws_base = wb[RAPORT_SHEET]
    else:
        ws_base = wb[wb.sheetnames[0]]

    # zmiana nazwy na 'raport' bez konfliktu
    if ws_base.title != RAPORT_SHEET:
        if RAPORT_SHEET in wb.sheetnames:
            wb.remove(wb[RAPORT_SHEET])
        ws_base.title = RAPORT_SHEET

    ws_raport = wb[RAPORT_SHEET]

    target_cols = REQ_COLS + VALUE_COLS
    final_header = _ensure_headers_on_report(ws_raport, target_cols)

    # UTWÓRZ / WYCZYŚĆ raport_odfiltrowane (same nagłówki)
    _ensure_headers_only_sheet(wb, RAPORT_ODF, final_header)

    wb.save(xlsx_path)
    print(f"[kolumny] OK -> {xlsx_path} | sheets={wb.sheetnames}")


def _gui_pick_and_add_columns():
    root = tk.Tk()
    root.withdraw()

    f = filedialog.askopenfilename(
        title="Wybierz plik raportu",
        filetypes=[("Pliki Excel", "*.xlsx *.xlsm")]
    )
    if not f:
        root.destroy()
        return

    try:
        ensure_report_columns(f)
        messagebox.showinfo(
            "kolumny.py",
            "OK: uzupełniono 'raport' i przygotowano 'raport_odfiltrowane' (same nagłówki)."
        )
    except Exception as e:
        messagebox.showerror("Błąd", str(e))

    root.destroy()


def parse_args():
    ap = argparse.ArgumentParser()
    ap.add_argument("--in", dest="inp", help="Plik raportu do modyfikacji")
    return ap.parse_args()


def main():
    args = parse_args()
    if args.inp:
        ensure_report_columns(args.inp)
        return
    _gui_pick_and_add_columns()


if __name__ == "__main__":
    if len(sys.argv) == 1:
        _gui_pick_and_add_columns()
    else:
        main()
